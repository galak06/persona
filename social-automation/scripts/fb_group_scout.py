"""
Facebook Group Scout — CLI version using Playwright.
Searches Facebook for new dog-related groups (public + private) to join.
Uses the same saved session as fb_scan.py (facebook_session.json).

Usage:
    python scripts/fb_group_scout.py           # normal run (weekly cadence)
    python scripts/fb_group_scout.py --force   # override weekly re-run guard
    python scripts/fb_group_scout.py --dry-run # score and list groups, no join requests
"""

from __future__ import annotations

import json
import re
import sys
import time
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "lib"))

from notifier import skill_started, skill_finished, skill_skipped

SESSION_FILE = PROJECT_ROOT / ".claude/state/facebook_session.json"
LAST_RUN_FILE = PROJECT_ROOT / ".claude/state/last_run.json"
PENDING_FILE = PROJECT_ROOT / ".claude/state/pending_groups.json"
LOG_FILE = PROJECT_ROOT / "logs/engagement_log.jsonl"
TRACKER_FILE = PROJECT_ROOT.parent.parent / "facebook_groups_tracker.xlsx"
ERROR_LOG = PROJECT_ROOT / "logs/errors.log"

JOIN_LIMIT_PER_WEEK = 3
MIN_SCORE = 40

# Niche keywords for scoring
FOOD_KW = ["food", "nutrition", "recipe", "diet", "raw", "kibble", "homemade",
           "feeding", "meal", "ingredient", "protein", "grain free"]
GPS_KW = ["gps", "tracker", "tracking", "running", "canicross", "hike",
          "hiking", "trail", "sport", "active", "agility"]
LIFESTYLE_KW = ["dog owner", "dog lifestyle", "dog product", "dog health",
                "dog care", "dog community", "dog lover"]
COMPETITOR_BRANDS = {
    "tractive", "fi collar", "ficollar", "whistle", "link akc",
    "ollie", "nom nom", "farmer's dog", "open farm",
}

SEARCH_QUERIES = [
    "homemade dog food",
    "raw dog food",
    "dog nutrition",
    "dog food recipes",
    "dog diet advice",
    "running with dogs",
    "canicross",
    "GPS dog tracker",
    "dog hiking",
    "dog owners community",
    "healthy dogs",
    "dog product reviews",
]

EXTRACT_GROUP_CARDS_JS = """
() => {
    const cards = [];
    const links = Array.from(document.querySelectorAll('a[href*="/groups/"]'));
    const seen = new Set();

    for (const a of links) {
        const href = a.getAttribute('href') || '';
        const match = href.match(/\/groups\/([^/?#]+)/);
        if (!match) continue;
        const gid = match[1];
        if (['feed', 'discover', 'search', 'explore', 'create'].includes(gid)) continue;
        if (seen.has(gid)) continue;
        seen.add(gid);

        // Walk up to find the card container (up to 8 levels)
        let card = a;
        for (let i = 0; i < 8; i++) {
            if (!card.parentElement) break;
            card = card.parentElement;
            if (card.children.length > 2) break;
        }

        const text = card.innerText || '';
        const lines = text.split('\\n').map(l => l.trim()).filter(Boolean);

        let memberText = '';
        let privacyText = 'public';
        let postFreq = '';
        let descLines = [];

        for (const line of lines) {
            const ll = line.toLowerCase();
            if (ll.match(/\\d.*member/)) memberText = line;
            else if (ll.includes('private')) privacyText = 'private';
            else if (ll.includes('public')) privacyText = 'public';
            else if (ll.match(/post/)) postFreq = line;
            else if (line.length > 20) descLines.push(line);
        }

        cards.push({
            url: 'https://www.facebook.com' + href.split('?')[0],
            name: lines[0] || gid,
            privacy: privacyText,
            member_text: memberText,
            post_frequency: postFreq,
            description: descLines.slice(0, 3).join(' '),
        });
    }
    return cards.slice(0, 25);
}
"""

FIND_JOIN_BUTTON_JS = """
() => {
    const candidates = Array.from(
        document.querySelectorAll('[role="button"], button')
    );
    for (const btn of candidates) {
        const label = (
            btn.getAttribute('aria-label') ||
            btn.innerText ||
            ''
        ).toLowerCase().trim();
        if (label === 'join group' || label === 'join' ||
            label === 'request to join' || label === 'request') {
            btn.click();
            return 'clicked:' + label;
        }
    }
    // Check if already joined / pending
    for (const btn of candidates) {
        const label = (btn.getAttribute('aria-label') || btn.innerText || '').toLowerCase();
        if (label.includes('joined') || label.includes('member')) return 'already_joined';
        if (label.includes('pending') || label.includes('requested')) return 'already_pending';
    }
    return 'not_found';
}
"""


def log_error(msg: str) -> None:
    ERROR_LOG.parent.mkdir(parents=True, exist_ok=True)
    ts = datetime.now(timezone.utc).isoformat()
    with ERROR_LOG.open("a") as f:
        f.write(f"[{ts}] [fb_group_scout] {msg}\n")


def load_last_run() -> dict:
    if LAST_RUN_FILE.exists():
        with LAST_RUN_FILE.open() as f:
            return json.load(f)
    return {}


def save_last_run(data: dict) -> None:
    LAST_RUN_FILE.parent.mkdir(parents=True, exist_ok=True)
    with LAST_RUN_FILE.open("w") as f:
        json.dump(data, f, indent=2)


def join_requests_this_week() -> int:
    """Count group_join_request log entries from the last 7 days."""
    if not LOG_FILE.exists():
        return 0
    week_ago = (date.today() - timedelta(days=7)).isoformat()
    count = 0
    with LOG_FILE.open() as f:
        for line in f:
            try:
                entry = json.loads(line)
                if (entry.get("action") == "group_join_request"
                        and entry.get("date", "") >= week_ago):
                    count += 1
            except Exception:
                continue
    return count


def load_known_groups() -> set[str]:
    """Return a set of known group URLs + lowercase names from engagement log."""
    known: set[str] = set()
    if LOG_FILE.exists():
        with LOG_FILE.open() as f:
            for line in f:
                try:
                    entry = json.loads(line)
                    if entry.get("action") == "group_join_request":
                        u = entry.get("target_url", "").lower()
                        n = entry.get("target_name", "").lower()
                        if u:
                            known.add(u)
                        if n:
                            known.add(n)
                except Exception:
                    continue
    # Also try to read tracker xlsx if it exists
    if TRACKER_FILE.exists():
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(TRACKER_FILE), read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value).lower().strip() if c.value else "" for c in next(ws.iter_rows())]
            url_col = next((i for i, h in enumerate(headers) if "url" in h), None)
            name_col = next((i for i, h in enumerate(headers) if "name" in h), None)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if url_col is not None and row[url_col]:
                    known.add(str(row[url_col]).lower())
                if name_col is not None and row[name_col]:
                    known.add(str(row[name_col]).lower())
            wb.close()
        except Exception:
            pass
    return known


def load_pending() -> list[dict]:
    """Load previously found groups that haven't been joined yet."""
    if PENDING_FILE.exists():
        try:
            return json.loads(PENDING_FILE.read_text())
        except Exception:
            pass
    return []


def save_pending(groups: list[dict]) -> None:
    """Persist the pending groups queue."""
    PENDING_FILE.parent.mkdir(parents=True, exist_ok=True)
    PENDING_FILE.write_text(json.dumps(groups, indent=2))


def add_to_pending(candidates: list[dict], known_groups: set[str]) -> int:
    """
    Merge new candidates into the pending queue.
    Skips groups already known (joined/requested) or already in pending.
    Returns count of newly added entries.
    """
    existing = load_pending()
    existing_urls = {g["url"].lower() for g in existing}
    added = 0
    for g in candidates:
        url = g["url"].lower()
        if url not in known_groups and url not in existing_urls:
            g["added_to_pending"] = date.today().isoformat()
            existing.append(g)
            existing_urls.add(url)
            added += 1
    save_pending(existing)
    return added


def remove_from_pending(joined_urls: list[str]) -> None:
    """Remove groups from pending once they've been joined/requested."""
    pending = load_pending()
    joined_set = {u.lower() for u in joined_urls}
    pending = [g for g in pending if g["url"].lower() not in joined_set]
    save_pending(pending)


def parse_member_count(text: str) -> int:
    if not text:
        return 0
    text = text.lower().replace(",", "")
    m = re.search(r"([\d.]+)\s*k", text)
    if m:
        return int(float(m.group(1)) * 1000)
    m = re.search(r"([\d.]+)\s*m", text)
    if m:
        return int(float(m.group(1)) * 1_000_000)
    m = re.search(r"(\d+)", text)
    return int(m.group(1)) if m else 0


def score_group(g: dict) -> int:
    score = 0
    text = (g["name"] + " " + g["description"]).lower()

    # Niche keyword match (additive, max 30)
    if any(kw in text for kw in FOOD_KW):
        score += 15
    if any(kw in text for kw in GPS_KW):
        score += 10
    if any(kw in text for kw in LIFESTYLE_KW):
        score += 5

    # Member count (max 20)
    mc = g["member_count"]
    if 1_000 <= mc <= 10_000:
        score += 20
    elif 10_000 < mc <= 50_000:
        score += 15
    elif 50_000 < mc <= 150_000:
        score += 10

    # Activity level (max 20)
    freq = g["post_frequency"].lower()
    if "day" in freq:
        score += 20
    elif "week" in freq:
        score += 10

    # Private group bonus
    if g["privacy"] == "private":
        score += 10

    # Competitor penalty
    if any(brand in text for brand in COMPETITOR_BRANDS):
        score -= 40

    return max(0, score)


def append_to_tracker(group: dict) -> None:
    """Append a new row to facebook_groups_tracker.xlsx."""
    if not TRACKER_FILE.exists():
        print(f"  [tracker] Tracker file not found at {TRACKER_FILE} — skipping xlsx update.")
        return
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(TRACKER_FILE))
        ws = wb.active
        # Find last row
        last_row = ws.max_row + 1
        ws.cell(last_row, 1, group["name"])
        ws.cell(last_row, 2, group["url"])
        ws.cell(last_row, 3, group["privacy"])
        ws.cell(last_row, 4, group["member_count"])
        ws.cell(last_row, 5, group["score"])
        ws.cell(last_row, 6, date.today().isoformat())
        ws.cell(last_row, 7, "join_requested")
        ws.cell(last_row, 8, group["found_via_query"])
        wb.save(str(TRACKER_FILE))
        print(f"  [tracker] Added row: {group['name']}")
    except Exception as e:
        print(f"  [tracker] WARNING: Could not update xlsx: {e}")


def log_join_request(group: dict, status: str) -> None:
    entry = {
        "date": date.today().isoformat(),
        "timestamp": datetime.now(timezone.utc).isoformat() + "Z",
        "platform": "facebook",
        "action": "group_join_request",
        "target_name": group["name"],
        "target_url": group["url"],
        "privacy": group["privacy"],
        "member_count": group["member_count"],
        "score": group["score"],
        "found_via": group["found_via_query"],
        "status": status,
    }
    LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
    with LOG_FILE.open("a") as f:
        f.write(json.dumps(entry) + "\n")


def get_user_approval(candidates: list[dict], budget: int) -> list[dict]:
    """
    Present scored group candidates to user and get approval.
    Returns list of approved groups.
    """
    print("\n" + "=" * 60)
    print(f"Facebook Group Scout — {len(candidates)} candidates found")
    print(f"Join budget this week: {budget}/3")
    print("=" * 60)

    for i, g in enumerate(candidates, 1):
        mc = f"{g['member_count']:,}" if g["member_count"] else "unknown"
        print(f"\n #{i}  {g['name']}  [{g['privacy'].upper()}]")
        print(f"      Members: {mc}  |  Score: {g['score']}  |  {g['post_frequency'] or 'activity unknown'}")
        print(f"      URL: {g['url']}")
        print(f"      Found via: \"{g['found_via_query']}\"")
        if g["description"]:
            desc = g["description"][:120]
            print(f"      Description: \"{desc}...\"")

    print("\n" + "-" * 60)
    print(f"Approve which groups to join/request? (max {budget} this week)")
    print("  Enter: 'all'  |  numbers like '1 3'  |  'none' to skip")
    response = input("Your choice: ").strip().lower()

    if response == "none" or not response:
        return []
    if response == "all":
        return candidates[:budget]

    # Parse numbers
    approved = []
    for token in response.split():
        try:
            idx = int(token) - 1
            if 0 <= idx < len(candidates):
                approved.append(candidates[idx])
        except ValueError:
            pass
    return approved[:budget]


def run_scout(dry_run: bool = False) -> None:
    from playwright.sync_api import sync_playwright
    import random

    print("=== Facebook Group Scout (CLI) ===\n")

    # Re-run guard (weekly)
    last_run = load_last_run()
    scout_last = last_run.get("fb_group_scout", {})
    scout_last_date = (scout_last.get("last_run_at") or "")[:10]
    if scout_last_date:
        days_since = (date.today() - date.fromisoformat(scout_last_date)).days
        if days_since < 7 and scout_last.get("status") == "success":
            print(f"SKIP: Already ran successfully {days_since} day(s) ago ({scout_last_date}).")
            print("Use --force to override.")
            skill_skipped("fb-group-scout", f"Already ran {days_since} day(s) ago")
            if "--force" not in sys.argv:
                return
            print("--force detected, re-running.\n")

    # Weekly budget check
    used_this_week = join_requests_this_week()
    budget = JOIN_LIMIT_PER_WEEK - used_this_week
    print(f"Join request budget: {budget}/3 remaining this week")
    if budget <= 0 and not dry_run:
        print("ABORT: Weekly join request limit reached (3/3). Try again next week.")
        skill_skipped("fb-group-scout", "Weekly join request limit reached (3/3)")
        return

    skill_started("fb-group-scout", f"Searching for new dog groups to join — budget: {budget}/3 this week")

    # Known groups (skip duplicates)
    known_groups = load_known_groups()
    print(f"Known groups to skip: {len(known_groups)}")

    # ── Check pending queue first ────────────────────────────────────────────
    pending = load_pending()
    # Filter out any that have since been joined
    pending = [g for g in pending if g["url"].lower() not in known_groups]
    save_pending(pending)

    if pending:
        print(f"\n📋 Pending queue: {len(pending)} group(s) found in previous runs, not yet joined")
        for i, g in enumerate(pending, 1):
            mc = f"{g['member_count']:,}" if g.get("member_count") else "?"
            added = g.get("added_to_pending", "?")
            print(
                f"  #{i:2} [{g['privacy'].upper():7}]  score={g['score']:3}  "
                f"members={mc:>8}  {g['name']}  (found {added})"
            )

        if not dry_run:
            print(f"\nJoin from pending queue first? (yes/no — budget: {budget}/3 this week)")
            resp = input("Your choice: ").strip().lower()
            if resp == "yes":
                to_join = pending[:budget]
                # Reuse the join logic below by injecting into candidates flow
                # We'll handle this in the join section
            else:
                pending = []  # user chose to skip pending, search fresh

    if not SESSION_FILE.exists():
        print("ERROR: No saved Facebook session found.")
        print("Run this first:  python scripts/fb_login.py")
        return

    # Collect + deduplicate candidates across all queries
    all_candidates: dict[str, dict] = {}  # url → group dict

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context(
            storage_state=str(SESSION_FILE),
            viewport={"width": 1280, "height": 900},
            user_agent=(
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/131.0.0.0 Safari/537.36"
            ),
        )
        page = context.new_page()

        # Quick session check
        print("\nChecking Facebook session...")
        page.goto("https://www.facebook.com/", wait_until="domcontentloaded")
        time.sleep(3)
        if "login" in page.url.lower():
            print("ABORT: Facebook session expired. Re-run fb_login.py")
            log_error("SESSION_EXPIRED")
            browser.close()
            return
        print("Facebook session OK.\n")

        # ── Search loop ──────────────────────────────────────────────────────
        for query in SEARCH_QUERIES:
            search_url = (
                f"https://www.facebook.com/search/groups/"
                f"?q={query.replace(' ', '%20')}"
            )
            print(f"Searching: \"{query}\"")
            try:
                page.goto(search_url, wait_until="domcontentloaded")
                time.sleep(4)

                # Scroll to load more results
                for _ in range(2):
                    page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                    time.sleep(2)

                raw_cards = page.evaluate(EXTRACT_GROUP_CARDS_JS)
                print(f"  Raw results: {len(raw_cards)}")

                for card in raw_cards:
                    url = card["url"].lower()

                    # Skip known groups
                    if url in known_groups or card["name"].lower() in known_groups:
                        continue

                    # Parse member count
                    card["member_count"] = parse_member_count(card["member_text"])
                    card["found_via_query"] = query
                    card["score"] = score_group(card)

                    # Skip below minimum score or out-of-range member count
                    if card["score"] < MIN_SCORE:
                        continue
                    mc = card["member_count"]
                    if mc > 0 and (mc < 1_000 or mc > 150_000):
                        continue

                    # Deduplicate across queries (keep highest score)
                    if url not in all_candidates or card["score"] > all_candidates[url]["score"]:
                        all_candidates[url] = card

            except Exception as e:
                print(f"  ERROR: {e}")
                log_error(f"SEARCH_FAILED: query='{query}' — {e}")

            time.sleep(random.uniform(3, 6))

        # ── Present shortlist ────────────────────────────────────────────────
        candidates = sorted(
            all_candidates.values(),
            key=lambda g: g["score"],
            reverse=True,
        )[:10]

        print(f"\nTotal qualifying candidates: {len(candidates)}")

        if not candidates:
            print("No new qualifying groups found. Done.")
            browser.close()
            # Still write last_run so weekly guard activates
            last_run["fb_group_scout"] = {
                "last_run_at": datetime.now(timezone.utc).isoformat(),
                "groups_found": 0,
                "groups_approved": 0,
                "join_requests_sent": 0,
                "status": "success",
            }
            save_last_run(last_run)
            return

        if dry_run:
            print("\n[DRY RUN — saving to pending queue, no join requests sent]\n")
            for i, g in enumerate(candidates, 1):
                mc = f"{g['member_count']:,}" if g.get("member_count") else "?"
                print(
                    f"  #{i:2} [{g['privacy'].upper():7}]  score={g['score']:3}  "
                    f"members={mc:>8}  {g['name']}"
                )
            added = add_to_pending(candidates, known_groups)
            print(f"\n✅ Saved {added} new group(s) to pending queue ({PENDING_FILE.name})")
            print(f"   Run without --dry-run to join them (budget: {budget}/3 this week)")
            browser.close()
            return

        # ── Merge pending into candidate pool ────────────────────────────────
        # Inject pending groups at the top (they were previously scored/approved)
        pending_to_join = [g for g in pending if g["url"].lower() not in known_groups]
        if pending_to_join:
            # Prepend pending to candidates, deduplicate by URL
            seen_urls = {g["url"].lower() for g in pending_to_join}
            fresh = [g for g in candidates if g["url"].lower() not in seen_urls]
            combined = pending_to_join + fresh
        else:
            combined = candidates

        # ── User approval ────────────────────────────────────────────────────
        approved = get_user_approval(combined, budget)

        if not approved:
            print("\nNo groups approved.")
            # Still save unapproved candidates to pending for next time
            added = add_to_pending(candidates, known_groups)
            print(f"Saved {added} candidate(s) to pending queue for next run.")
            browser.close()
            return

        # ── Send join requests ───────────────────────────────────────────────
        join_requests_sent = 0
        print(f"\nSending {len(approved)} join request(s)...\n")

        for group in approved:
            print(f"  → {group['name']} [{group['privacy'].upper()}]")
            try:
                page.goto(group["url"], wait_until="domcontentloaded")
                time.sleep(4)

                result = page.evaluate(FIND_JOIN_BUTTON_JS)
                print(f"     Button result: {result}")

                if result.startswith("clicked"):
                    time.sleep(2)
                    status = "join_requested" if group["privacy"] == "private" else "joined"
                    log_join_request(group, status)
                    append_to_tracker(group)
                    join_requests_sent += 1
                    known_groups.add(group["url"].lower())
                    label = "✅ Request sent (pending admin approval)" if group["privacy"] == "private" else "✅ Joined immediately"
                    print(f"     {label}")
                elif result == "already_joined":
                    print("     SKIP: Already a member")
                elif result == "already_pending":
                    print("     SKIP: Request already pending")
                else:
                    log_error(f"JOIN_BUTTON_NOT_FOUND: {group['url']}")
                    print("     WARNING: Join button not found — check manually")

            except Exception as e:
                msg = f"JOIN_FAILED: {group['name']} — {e}"
                print(f"     ERROR: {e}")
                log_error(msg)

            # Random delay between requests (60–180s)
            if group != approved[-1]:
                delay = random.uniform(60, 180)
                print(f"     Waiting {delay:.0f}s before next request...")
                time.sleep(delay)

        # Save refreshed session
        context.storage_state(path=str(SESSION_FILE))
        browser.close()

    # Remove joined groups from pending queue
    joined_urls = [g["url"] for g in approved if g in approved]
    remove_from_pending(joined_urls)

    # Save any approved-but-not-joined (budget ran out mid-approval) to pending
    not_joined = [g for g in approved[join_requests_sent:] if join_requests_sent < len(approved)]
    if not_joined:
        added = add_to_pending(not_joined, known_groups)
        print(f"Budget exhausted — saved {added} remaining approved group(s) to pending queue.")

    # Save all new candidates not yet joined to pending for future runs
    unapproved = [g for g in candidates if g not in approved]
    added = add_to_pending(unapproved, known_groups)
    if added:
        print(f"Saved {added} unapproved candidate(s) to pending queue for next run.")

    # Update last run
    last_run["fb_group_scout"] = {
        "last_run_at": datetime.now(timezone.utc).isoformat(),
        "groups_found": len(candidates),
        "groups_approved": len(approved),
        "join_requests_sent": join_requests_sent,
        "status": "success",
    }
    save_last_run(last_run)

    # Summary
    pending_remaining = len(load_pending())
    print(f"""
=== Facebook Group Scout Complete ===
Queries searched:       {len(SEARCH_QUERIES)}
Candidates evaluated:   {len(all_candidates)}
Surfaced (score ≥{MIN_SCORE}):   {len(candidates)}
Approved by you:        {len(approved)}
Join requests sent:     {join_requests_sent}
Budget remaining:       {budget - join_requests_sent}/3 this week
Pending queue:          {pending_remaining} group(s) saved for next run
""")
    summary = (
        f"🔍 Searched {len(SEARCH_QUERIES)} queries, found {len(candidates)} candidates\n"
        f"✅ Join requests sent: {join_requests_sent} "
        f"(budget remaining: {budget - join_requests_sent}/3 this week)"
    )
    skill_finished("fb-group-scout", summary)


if __name__ == "__main__":
    dry_run = "--dry-run" in sys.argv
    run_scout(dry_run=dry_run)
