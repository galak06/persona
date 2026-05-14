"""State I/O + logging for fb-group-scout."""

from __future__ import annotations

import json
from datetime import UTC, date, datetime, timedelta
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parents[2]

SESSION_FILE = PROJECT_ROOT / ".claude/state/facebook_session.json"
LAST_RUN_FILE = PROJECT_ROOT / ".claude/state/last_run.json"
PENDING_FILE = PROJECT_ROOT / ".claude/state/pending_groups.json"
LOG_FILE = PROJECT_ROOT / "logs/engagement_log.jsonl"
TRACKER_FILE = PROJECT_ROOT.parent.parent / "facebook_groups_tracker.xlsx"
JSON_TRACKER_FILE = PROJECT_ROOT / "data/groups_tracker.json"
ERROR_LOG = PROJECT_ROOT / "logs/errors.log"


def log_error(msg: str) -> None:
    ERROR_LOG.parent.mkdir(parents=True, exist_ok=True)
    ts = datetime.now(UTC).isoformat()
    with ERROR_LOG.open("a") as f:
        f.write(f"[{ts}] [fb_group_scout] {msg}\n")


def load_last_run() -> dict:
    if LAST_RUN_FILE.exists():
        with LAST_RUN_FILE.open() as f:
            return json.load(f)
    return {}


def save_last_run(data: dict) -> None:
    LAST_RUN_FILE.parent.mkdir(parents=True, exist_ok=True)
    with LAST_RUN_FILE.open("w") as f:
        json.dump(data, f, indent=2)


def _count_join_requests_since(cutoff_iso_date: str) -> int:
    if not LOG_FILE.exists():
        return 0
    count = 0
    with LOG_FILE.open() as f:
        for line in f:
            try:
                entry = json.loads(line)
                if (
                    entry.get("action") == "group_join_request"
                    and entry.get("date", "") >= cutoff_iso_date
                ):
                    count += 1
            except Exception:
                continue
    return count


def join_requests_this_week() -> int:
    week_ago = (date.today() - timedelta(days=7)).isoformat()
    return _count_join_requests_since(week_ago)


def join_requests_today() -> int:
    return _count_join_requests_since(date.today().isoformat())


def load_known_groups() -> set[str]:
    """Return a set of known group URLs + lowercase names from engagement log + tracker."""
    known: set[str] = set()
    if LOG_FILE.exists():
        with LOG_FILE.open() as f:
            for line in f:
                try:
                    entry = json.loads(line)
                    if entry.get("action") == "group_join_request":
                        u = entry.get("target_url", "").lower()
                        n = entry.get("target_name", "").lower()
                        if u:
                            known.add(u)
                        if n:
                            known.add(n)
                except Exception:
                    continue
    if TRACKER_FILE.exists():
        try:
            import openpyxl

            wb = openpyxl.load_workbook(str(TRACKER_FILE), read_only=True, data_only=True)
            ws = wb.active
            headers = [
                str(c.value).lower().strip() if c.value else "" for c in next(ws.iter_rows())
            ]
            url_col = next((i for i, h in enumerate(headers) if "url" in h), None)
            name_col = next((i for i, h in enumerate(headers) if "name" in h), None)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if url_col is not None and row[url_col]:
                    known.add(str(row[url_col]).lower())
                if name_col is not None and row[name_col]:
                    known.add(str(row[name_col]).lower())
            wb.close()
        except Exception:
            pass
    return known


def load_pending() -> list[dict]:
    if PENDING_FILE.exists():
        try:
            return json.loads(PENDING_FILE.read_text())
        except Exception:
            pass
    return []


def save_pending(groups: list[dict]) -> None:
    PENDING_FILE.parent.mkdir(parents=True, exist_ok=True)
    PENDING_FILE.write_text(json.dumps(groups, indent=2))


def add_to_pending(candidates: list[dict], known_groups: set[str]) -> int:
    existing = load_pending()
    existing_urls = {g["url"].lower() for g in existing}
    added = 0
    for g in candidates:
        url = g["url"].lower()
        if url not in known_groups and url not in existing_urls:
            g["added_to_pending"] = date.today().isoformat()
            existing.append(g)
            existing_urls.add(url)
            added += 1
    save_pending(existing)
    return added


def remove_from_pending(joined_urls: list[str]) -> None:
    pending = load_pending()
    joined_set = {u.lower() for u in joined_urls}
    pending = [g for g in pending if g["url"].lower() not in joined_set]
    save_pending(pending)


def _now_iso_z() -> str:
    """UTC now in ISO 8601 with trailing Z, matching fb_notification_scan format."""
    return datetime.now(UTC).isoformat().replace("+00:00", "Z")


def _upsert_json_tracker(group: dict, status: str) -> str:
    """Upsert into data/groups_tracker.json. status is 'joined' or 'join_requested'.

    Public groups (status='joined') get joined_at=now so warmup gates can tick.
    Private groups (status='join_requested') get no joined_at — fb_notification_scan
    will fill it when the admin approves. Idempotent: never overwrites an existing
    joined_at on re-runs.

    Returns: 'added', 'updated', or 'unchanged'.
    """
    JSON_TRACKER_FILE.parent.mkdir(parents=True, exist_ok=True)
    tracker: list[dict] = []
    if JSON_TRACKER_FILE.exists():
        try:
            tracker = json.loads(JSON_TRACKER_FILE.read_text())
        except json.JSONDecodeError:
            tracker = []

    url = group["url"]
    now = _now_iso_z()

    for existing in tracker:
        if existing.get("group_url") == url:
            # Only escalate to joined; never downgrade or overwrite joined_at.
            if status == "joined" and existing.get("status") != "joined":
                existing["status"] = "joined"
                existing.setdefault("joined_at", now)
                JSON_TRACKER_FILE.write_text(json.dumps(tracker, indent=2))
                return "updated"
            return "unchanged"

    entry: dict = {
        "group_name": group["name"],
        "group_url": url,
        "status": status,
        "rules": "unknown",
        "last_post_at": None,
        "source_notification": f"fb_group_scout ({group.get('privacy', 'unknown')})",
        "privacy": group.get("privacy", "unknown"),
        "member_count": group.get("member_count"),
    }
    if status == "joined":
        entry["joined_at"] = now
    tracker.append(entry)
    JSON_TRACKER_FILE.write_text(json.dumps(tracker, indent=2))
    return "added"


def append_to_tracker(group: dict, status: str = "join_requested") -> None:
    """Record a join action.

    Writes the canonical JSON tracker (data/groups_tracker.json) — the source
    used by fb_group_post + warmup gates. Also best-effort updates the legacy
    xlsx tracker if present.
    """
    action = _upsert_json_tracker(group, status)
    print(f"  [tracker] {action}: {group['name']} (status={status})")

    if not TRACKER_FILE.exists():
        return
    try:
        import openpyxl

        wb = openpyxl.load_workbook(str(TRACKER_FILE))
        ws = wb.active
        last_row = ws.max_row + 1
        ws.cell(last_row, 1, group["name"])
        ws.cell(last_row, 2, group["url"])
        ws.cell(last_row, 3, group["privacy"])
        ws.cell(last_row, 4, group["member_count"])
        ws.cell(last_row, 5, group["score"])
        ws.cell(last_row, 6, date.today().isoformat())
        ws.cell(last_row, 7, status)
        ws.cell(last_row, 8, group["found_via_query"])
        wb.save(str(TRACKER_FILE))
    except Exception as e:
        print(f"  [tracker] WARNING: Could not update xlsx: {e}")


def log_join_request(group: dict, status: str) -> None:
    entry = {
        "date": date.today().isoformat(),
        "timestamp": datetime.now(UTC).isoformat() + "Z",
        "platform": "facebook",
        "action": "group_join_request",
        "target_name": group["name"],
        "target_url": group["url"],
        "privacy": group["privacy"],
        "member_count": group["member_count"],
        "score": group["score"],
        "found_via": group["found_via_query"],
        "competitor_mentions": group.get("competitor_mentions", 0),
        "status": status,
    }
    LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
    with LOG_FILE.open("a") as f:
        f.write(json.dumps(entry) + "\n")
